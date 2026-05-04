"""Contract tests for ``extract_animals_block`` (yearly format).

Each test builds a synthetic worksheet through ``openpyxl.Workbook`` —
no disk fixtures, deterministic, fast.

Geometry mirrors production yearly file ``raw_data/2025_рік.xlsx``:

  row 1: ignored (banner)
  row 2: ignored
  row 3: year labels — col B='2022 рік', col C='2023 рік'
  row 4: section header — col A='Чисельність/кількість лімітів '
  row 5+: species rows, each cell = '<species_alias> <population>/<limit>'
"""
from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from etl.models import AnimalValue
from etl.parsers_animals import (
    extract_animals_block,
    extract_animals_block_osnovni,
)
from etl.report_metadata import ReportMetadata


def _meta(
    *,
    report_type: str = "operational",
    priority: int = 10,
    vintage: datetime | None = None,
) -> ReportMetadata:
    return ReportMetadata(
        report_type=report_type,  # type: ignore[arg-type]
        vintage_date=vintage or datetime(2026, 1, 31),
        source_priority=priority,
    )


def _build_ws(
    species_rows: list[list[str | None]],
    *,
    year_labels: list[str | None] | None = None,
    extra_rows_after: list[list[Any]] | None = None,
) -> tuple[Worksheet, int]:
    """Build a synthetic animals sheet.

    species_rows: list of rows, each a list of cell values starting at
                  col B (col A is empty for animals body).
    year_labels:  values for row 3, cols B..N (default
                  ['2022 рік', '2023 рік']).
    extra_rows_after: optional rows injected after species (col A first).

    Returns (ws, header_row) — header_row is always 4.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "2025"

    # Year labels at row 3
    labels = year_labels if year_labels is not None else ["2022 рік", "2023 рік"]
    for offset, label in enumerate(labels):
        if label is not None:
            ws.cell(3, 2 + offset).value = label

    # Section header at row 4
    header_row = 4
    ws.cell(header_row, 1).value = "Чисельність/кількість лімітів "

    # Species rows starting at row 5
    next_row = header_row + 1
    for row_cells in species_rows:
        for offset, cell in enumerate(row_cells):
            if cell is not None:
                ws.cell(next_row, 2 + offset).value = cell
        next_row += 1

    for extra in extra_rows_after or []:
        for offset, cell in enumerate(extra):
            if cell is not None:
                ws.cell(next_row, 1 + offset).value = cell
        next_row += 1

    return ws, header_row


def _extract(
    ws: Worksheet,
    header_row: int,
) -> tuple[list[AnimalValue], list[str]]:
    return extract_animals_block(
        ws,
        source_file="synthetic.xlsx",
        base_meta=_meta(),
        header_row=header_row,
    )


# ---------------------------------------------------------------------------
# 1. Happy path — 6 species × 2 years = 12 emits
# ---------------------------------------------------------------------------

def test_happy_path_two_years() -> None:
    """Mirror of the 2025_рік.xlsx animals geometry: 6 species, 2 years."""
    ws, hdr = _build_ws(
        [
            ["Олень благор. 3787/*",   "Олень благор. 3697/*"],
            ["Олень плямистий 1025/*", "Олень плямистий 744/*"],
            ["Козуля 35191/*",         "Козуля 30810/*"],
            ["Кабан 7700/*",           "Кабан 6813/*"],
            ["Лань 362/*",             "Лань 334/*"],
            ["Муфлон 302/*",           "Муфлон 312/*"],
        ],
    )
    animals, warnings = _extract(ws, hdr)

    assert len(animals) == 12
    assert warnings == []
    species_seen = {a.species_name for a in animals}
    years_seen = sorted({a.year for a in animals})
    assert years_seen == [2022, 2023]
    assert species_seen == {
        "Олень благор.", "Олень плямистий", "Козуля",
        "Кабан", "Лань", "Муфлон",
    }
    # Spot-check one emit (Олень благор. 2022)
    deer_2022 = next(
        a for a in animals
        if a.species_name == "Олень благор." and a.year == 2022
    )
    assert deer_2022.population == 3787
    assert deer_2022.limit_qty is None
    assert deer_2022.raw_text == "Олень благор. 3787/*"


# ---------------------------------------------------------------------------
# 2. Limit '*' becomes None
# ---------------------------------------------------------------------------

def test_limit_star_becomes_none() -> None:
    """``"*"`` in the limit slot maps to ``limit_qty=None``."""
    ws, hdr = _build_ws([["Кабан 7700/*", None]])
    animals, _ = _extract(ws, hdr)
    assert len(animals) == 1
    assert animals[0].limit_qty is None
    assert animals[0].population == 7700


# ---------------------------------------------------------------------------
# 3. Numeric limit parsed as int
# ---------------------------------------------------------------------------

def test_limit_numeric_parsed_as_int() -> None:
    """Future-shape: when the limit slot carries a number, it is stored."""
    ws, hdr = _build_ws([["Кабан 7700/250", None]])
    animals, _ = _extract(ws, hdr)
    assert len(animals) == 1
    assert animals[0].limit_qty == 250
    assert animals[0].population == 7700


# ---------------------------------------------------------------------------
# 4. Abbreviated species name preserved
# ---------------------------------------------------------------------------

def test_abbreviated_species_name_preserved() -> None:
    """``"Олень благор."`` (with trailing period) kept verbatim."""
    ws, hdr = _build_ws([["Олень благор. 3787/*", None]])
    animals, _ = _extract(ws, hdr)
    assert len(animals) == 1
    assert animals[0].species_name == "Олень благор."


# ---------------------------------------------------------------------------
# 5. Double space between name and population
# ---------------------------------------------------------------------------

def test_double_space_in_cell() -> None:
    """``"Олень плямистий  650/*"`` (production variant) parses cleanly."""
    ws, hdr = _build_ws([["Олень плямистий  650/*", None]])
    animals, _ = _extract(ws, hdr)
    assert len(animals) == 1
    assert animals[0].species_name == "Олень плямистий"
    assert animals[0].population == 650


# ---------------------------------------------------------------------------
# 6. raw_text preserved verbatim
# ---------------------------------------------------------------------------

def test_raw_text_preserved() -> None:
    """``raw_text`` is the original cell content, no normalization."""
    raw = "Олень плямистий  650/*"  # double space — production variant
    ws, hdr = _build_ws([[raw, None]])
    animals, _ = _extract(ws, hdr)
    assert len(animals) == 1
    assert animals[0].raw_text == raw


# ---------------------------------------------------------------------------
# 7. Section header not found → empty result, no warnings
# ---------------------------------------------------------------------------

def test_section_header_not_found() -> None:
    """Worksheet without the section header → ``([], [])``."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(1, 1).value = "Some other content"
    ws.cell(2, 1).value = "Загальна реалізація, млн. грн"

    animals, warnings = extract_animals_block(
        ws,
        source_file="synthetic.xlsx",
        base_meta=_meta(),
    )
    assert animals == []
    assert warnings == []


# ---------------------------------------------------------------------------
# 8. Year labels missing → warning, no emits
# ---------------------------------------------------------------------------

def test_year_labels_missing() -> None:
    """Header found but row above is empty → info-level warning."""
    ws, hdr = _build_ws(
        [["Кабан 7700/*", None]],
        year_labels=[None, None],  # row 3 empty
    )
    animals, warnings = _extract(ws, hdr)

    assert animals == []
    assert len(warnings) == 1
    assert "no year labels found" in warnings[0]


# ---------------------------------------------------------------------------
# Osnovni-format tests (sub-step 5.5.3)
# ---------------------------------------------------------------------------

# Production parser_osnovni.ANNUAL_COLUMNS shape: 2022..2025 closed,
# 2026 is_ytd=True.
_OSNOVNI_ANNUAL_COLUMNS: tuple[tuple[int, int, bool], ...] = (
    (2, 2022, False),
    (3, 2023, False),
    (4, 2024, False),
    (5, 2025, False),
    (6, 2026, True),
)


def _build_osnovni_ws(
    species_rows: list[list[str | None]],
    *,
    header_row_offset: int = 0,
) -> tuple[Worksheet, int]:
    """Build a synthetic osnovni-format animals sheet.

    species_rows: list of rows, each list of cell values starting at
                  col B (cols B..F = 2022..2026 per ANNUAL_COLUMNS).
    header_row_offset: 0 means species data lives in cols B+ of the
                  same row as the col-A section header (production
                  geometry). 1 means species data lives in the row
                  AFTER the header (alt synthetic geometry).

    Returns (ws, header_row).
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Основні показники"

    # Section header at row 4
    header_row = 4
    ws.cell(header_row, 1).value = "Чисельність/кількість лімітів "

    next_row = header_row if header_row_offset == 0 else header_row + 1
    for row_cells in species_rows:
        for offset, cell in enumerate(row_cells):
            if cell is not None:
                ws.cell(next_row, 2 + offset).value = cell
        next_row += 1

    return ws, header_row


def _extract_osnovni(
    ws: Worksheet,
    header_row: int,
    *,
    annual_columns: tuple[tuple[int, int, bool], ...] = _OSNOVNI_ANNUAL_COLUMNS,
) -> tuple[list[AnimalValue], list[str]]:
    return extract_animals_block_osnovni(
        ws,
        annual_columns=annual_columns,
        source_file="synthetic.xlsx",
        base_meta=_meta(),
        header_row=header_row,
    )


# ---------------------------------------------------------------------------
# 9. Happy path — 6 species × 4 closed years = 24 emits, YTD col skipped
# ---------------------------------------------------------------------------

def test_osnovni_happy_path_four_closed_years() -> None:
    """Mirror of the osnovni file's animals geometry: 6 species, 4 closed
    years (2022-2025). 2026 column is is_ytd=True and must be skipped."""
    ws, hdr = _build_osnovni_ws(
        [
            # cols B (2022), C (2023), D (2024), E (2025), F (2026 YTD)
            ["Олень благор. 3787/*",   "Олень благор. 3697/*",
             "Олень благор. 3678/*",   "Олень благор. 3644/*",   None],
            ["Олень плямистий 1025/*", "Олень плямистий 744/*",
             "Олень плямистий 895/*",  "Олень плямистий 650/*",  None],
            ["Козуля 35191/*",         "Козуля 30810/*",
             "Козуля 32381/*",         "Козуля 34024/*",         None],
            ["Кабан 7700/*",           "Кабан 6813/*",
             "Кабан 7516/*",           "Кабан 7562/*",           None],
            ["Лань 362/*",             "Лань 334/*",
             "Лань 127/*",             "Лань 164/*",             None],
            ["Муфлон 302/*",           "Муфлон 312/*",
             "Муфлон 112/*",           "Муфлон 126/*",           None],
        ],
    )
    animals, warnings = _extract_osnovni(ws, hdr)

    assert len(animals) == 24
    assert warnings == []
    years_seen = sorted({a.year for a in animals})
    assert years_seen == [2022, 2023, 2024, 2025]
    assert 2026 not in years_seen  # YTD col skipped
    species_seen = {a.species_name for a in animals}
    assert species_seen == {
        "Олень благор.", "Олень плямистий", "Козуля",
        "Кабан", "Лань", "Муфлон",
    }
    # Spot-check Олень благор. 2025 (matches production)
    deer_2025 = next(
        a for a in animals
        if a.species_name == "Олень благор." and a.year == 2025
    )
    assert deer_2025.population == 3644
    assert deer_2025.limit_qty is None


# ---------------------------------------------------------------------------
# 10. is_ytd column ignored even when populated
# ---------------------------------------------------------------------------

def test_osnovni_skips_ytd_column() -> None:
    """A populated is_ytd column must NOT produce emits — animal census
    only reflects closed years."""
    ws, hdr = _build_osnovni_ws(
        [
            # 2022..2025 closed + 2026 YTD populated (synthetic)
            ["Кабан 7700/*", "Кабан 6813/*", "Кабан 7516/*",
             "Кабан 7562/*", "Кабан 9999/*"],
        ],
    )
    animals, _ = _extract_osnovni(ws, hdr)

    assert len(animals) == 4  # 2022..2025
    years_seen = sorted({a.year for a in animals})
    assert years_seen == [2022, 2023, 2024, 2025]
    populations = sorted({a.population for a in animals})
    assert 9999 not in populations  # 2026 YTD value not emitted


# ---------------------------------------------------------------------------
# 11. Cells outside annual_columns ignored
# ---------------------------------------------------------------------------

def test_osnovni_ignores_columns_outside_annual_columns() -> None:
    """Cells in cols not listed in annual_columns (e.g., col 7 for
    current month) are simply not iterated — no warnings, no emits."""
    ws, hdr = _build_osnovni_ws(
        [
            # cols B-F as expected, plus col 7 stray data
            ["Кабан 7700/*", "Кабан 6813/*", "Кабан 7516/*",
             "Кабан 7562/*", None, "Кабан 1234/*"],
        ],
    )
    animals, warnings = _extract_osnovni(ws, hdr)

    # Only B-E (4 closed years) emit; col 7 untouched.
    assert len(animals) == 4
    assert all(a.population != 1234 for a in animals)
    assert warnings == []


# ---------------------------------------------------------------------------
# 12. Section header not found → empty + no warnings
# ---------------------------------------------------------------------------

def test_osnovni_section_header_not_found() -> None:
    """Worksheet without 'Чисельність/кількість лімітів' header → ([], [])."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(1, 1).value = "Some other content"

    animals, warnings = extract_animals_block_osnovni(
        ws,
        annual_columns=_OSNOVNI_ANNUAL_COLUMNS,
        source_file="synthetic.xlsx",
        base_meta=_meta(),
    )
    assert animals == []
    assert warnings == []


# ---------------------------------------------------------------------------
# 13. First species sits on the header row itself (production geometry)
# ---------------------------------------------------------------------------

def test_osnovni_first_species_on_header_row() -> None:
    """Production osnovni puts the first species data in cols B+ of the
    header row (col A has the section title). Scanner must include
    offset=0 — same fix as yearly parser."""
    ws, hdr = _build_osnovni_ws(
        [
            ["Олень благор. 3787/*", "Олень благор. 3697/*",
             "Олень благор. 3678/*", "Олень благор. 3644/*", None],
            ["Олень плямистий 1025/*", "Олень плямистий 744/*",
             "Олень плямистий 895/*",  "Олень плямистий 650/*", None],
            ["Козуля 35191/*", "Козуля 30810/*",
             "Козуля 32381/*", "Козуля 34024/*", None],
            ["Кабан 7700/*", "Кабан 6813/*",
             "Кабан 7516/*", "Кабан 7562/*", None],
            ["Лань 362/*", "Лань 334/*",
             "Лань 127/*", "Лань 164/*", None],
            ["Муфлон 302/*", "Муфлон 312/*",
             "Муфлон 112/*", "Муфлон 126/*", None],
        ],
        header_row_offset=0,  # data starts AT header_row, not below
    )
    animals, _ = _extract_osnovni(ws, hdr)

    assert len(animals) == 24  # 6 species × 4 closed years
    species_seen = {a.species_name for a in animals}
    assert "Олень благор." in species_seen


# ---------------------------------------------------------------------------
# 14. All annual_columns marked is_ytd → empty + warning
# ---------------------------------------------------------------------------

def test_osnovni_all_columns_marked_ytd_returns_empty() -> None:
    """Edge case: every column is is_ytd=True — extractor returns empty
    with an info-level warning."""
    ws, hdr = _build_osnovni_ws(
        [["Кабан 7700/*", "Кабан 6813/*"]],
    )
    animals, warnings = _extract_osnovni(
        ws, hdr,
        annual_columns=((2, 2026, True), (3, 2027, True)),
    )
    assert animals == []
    assert len(warnings) == 1
    assert "no closed-year columns" in warnings[0]
