"""``etl/cli/writeback.py`` — CLI dry-run / commit dispatch + parser detection.

Network/DB code is mocked at module boundary (``open_connection``,
``PostgresRepository``). Real pipeline runs end-to-end on real Excel
fixtures from ``raw_data/`` so detect_parser is exercised against actual
on-disk workbooks; tests skip cleanly if those files aren't present.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from etl.cli.writeback import (
    SUMMARY_WARN_LIMIT,
    PipelineOutput,
    detect_parser,
    main,
)
from etl.db.code_mapper import UnknownMetricError
from etl.db.interface import WriteResult
from etl.models import ParseResult, ReferenceText

OSNOVNI_FIXTURE = Path("raw_data/Основні_показники_березень_2026_остання.xlsx")
ANNUAL_MONTHLY_FIXTURE = Path("raw_data/2025_рік.xlsx")


# ---------------------------------------------------------------------------
# detect_parser — runs against real workbooks
# ---------------------------------------------------------------------------

def test_detect_parser_dispatches_osnovni() -> None:
    if not OSNOVNI_FIXTURE.exists():
        pytest.skip(f"fixture not present: {OSNOVNI_FIXTURE}")
    assert detect_parser(OSNOVNI_FIXTURE) == "osnovni_annual"


def test_detect_parser_dispatches_annual_monthly() -> None:
    if not ANNUAL_MONTHLY_FIXTURE.exists():
        pytest.skip(f"fixture not present: {ANNUAL_MONTHLY_FIXTURE}")
    assert detect_parser(ANNUAL_MONTHLY_FIXTURE) == "annual_monthly"


def test_detect_parser_raises_for_unknown_format(tmp_path: Path) -> None:
    bad = tmp_path / "garbage.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    # Random text in B/C of every row near the top — neither year-headers
    # nor consecutive months. Both detection criteria must reject.
    for row in range(1, 6):
        ws.cell(row=row, column=2).value = "random text"
        ws.cell(row=row, column=3).value = "more random"
    wb.save(bad)

    with pytest.raises(ValueError) as excinfo:
        detect_parser(bad)
    msg = str(excinfo.value)
    assert "osnovni_annual rejected" in msg
    assert "annual_monthly rejected" in msg


# ---------------------------------------------------------------------------
# Helpers — synthesize a PipelineOutput so tests don't depend on real Excel
# parsing (faster + deterministic for the commit/dry-run paths)
# ---------------------------------------------------------------------------

def _fake_pipeline(
    *,
    warnings: list[str] | None = None,
    errors: list[str] | None = None,
) -> PipelineOutput:
    raw = ParseResult(warnings=list(warnings or []), errors=list(errors or []))
    return PipelineOutput(
        parser_name="annual_monthly",
        raw_result=raw,
        canonical_annual=[],
        canonical_monthly=[],
        canonical_species_annual=[],
        canonical_species_monthly=[],
        canonical_reference=[],
        canonical_salary=[],
        derived_annual=[],
    )


def _existing_xlsx(tmp_path: Path) -> Path:
    """Empty xlsx file just so ``main()``'s ``is_file()`` check passes."""
    p = tmp_path / "fake.xlsx"
    p.touch()
    return p


# ---------------------------------------------------------------------------
# Dry-run vs commit dispatch
# ---------------------------------------------------------------------------

def test_dry_run_does_not_open_connection(tmp_path: Path) -> None:
    fake_xlsx = _existing_xlsx(tmp_path)

    with (
        patch("etl.cli.writeback.run_full_pipeline", return_value=_fake_pipeline()),
        patch("etl.db.connection.open_connection") as mock_conn,
    ):
        rc = main([str(fake_xlsx)])

    assert rc == 0
    mock_conn.assert_not_called()


def test_commit_calls_write_batch(tmp_path: Path) -> None:
    fake_xlsx = _existing_xlsx(tmp_path)
    pipeline = _fake_pipeline()

    fake_outcome = WriteResult(
        batch_id=__import__("uuid").uuid4(),
        rows_to_revisions=0,
        rows_to_canonical=0,
        rows_unchanged=0,
        rows_superseded=0,
    )
    fake_repo = MagicMock()
    fake_repo.write_batch.return_value = fake_outcome

    cm = MagicMock()
    cm.__enter__.return_value = MagicMock()  # the conn
    cm.__exit__.return_value = False

    with (
        patch("etl.cli.writeback.run_full_pipeline", return_value=pipeline),
        patch("etl.db.connection.open_connection", return_value=cm) as mock_conn,
        patch("etl.db.postgres.PostgresRepository", return_value=fake_repo) as mock_cls,
    ):
        rc = main([str(fake_xlsx), "--commit"])

    assert rc == 0
    mock_conn.assert_called_once()
    mock_cls.assert_called_once()
    fake_repo.write_batch.assert_called_once()
    # Inspect the WriteBatch passed in.
    batch = fake_repo.write_batch.call_args.args[0]
    assert batch.source_file == str(fake_xlsx)
    # Empty pipeline → empty lists.
    assert batch.annual == []
    assert batch.monthly == []
    assert batch.species_annual == []
    assert batch.species_monthly == []


def test_unknown_metric_returns_exit_code_6(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    fake_xlsx = _existing_xlsx(tmp_path)

    fake_repo = MagicMock()
    fake_repo.write_batch.side_effect = UnknownMetricError(
        "Unknown Python metric_code: 'mystery_metric'."
    )

    cm = MagicMock()
    cm.__enter__.return_value = MagicMock()
    cm.__exit__.return_value = False

    with (
        patch("etl.cli.writeback.run_full_pipeline", return_value=_fake_pipeline()),
        patch("etl.db.connection.open_connection", return_value=cm),
        patch("etl.db.postgres.PostgresRepository", return_value=fake_repo),
    ):
        rc = main([str(fake_xlsx), "--commit"])

    assert rc == 6
    err = capsys.readouterr().err
    assert "unknown metric_code" in err.lower()


def test_missing_env_returns_exit_code_4(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    fake_xlsx = _existing_xlsx(tmp_path)

    with (
        patch("etl.cli.writeback.run_full_pipeline", return_value=_fake_pipeline()),
        patch(
            "etl.db.connection.open_connection",
            side_effect=FileNotFoundError("Environment file not found: .env"),
        ),
    ):
        rc = main([str(fake_xlsx), "--commit"])

    assert rc == 4
    err = capsys.readouterr().err
    assert "missing .env" in err.lower() or ".env" in err


# ---------------------------------------------------------------------------
# Stdout formatting
# ---------------------------------------------------------------------------

def test_summary_format_contains_key_fields(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    fake_xlsx = _existing_xlsx(tmp_path)
    with patch("etl.cli.writeback.run_full_pipeline", return_value=_fake_pipeline()):
        rc = main([str(fake_xlsx)])

    assert rc == 0
    out = capsys.readouterr().out
    for token in (
        "WRITEBACK PLAN",
        "Parser:",
        "Annual values:",
        "Monthly values:",
        "Species annual:",
        "Species monthly:",
        "Derived metrics:",
        "Period coverage:",
        "Warnings:",
        "Errors:",
        "[DRY-RUN]",
    ):
        assert token in out, f"missing expected token: {token!r}"


def test_verbose_full_warning_list(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    fake_xlsx = _existing_xlsx(tmp_path)
    # 8 warnings → top 5 appear in summary, verbose continues from #6.
    pipeline = _fake_pipeline(warnings=[f"warn-{i}" for i in range(8)])
    with patch("etl.cli.writeback.run_full_pipeline", return_value=pipeline):
        rc = main([str(fake_xlsx), "-v"])

    assert rc == 0
    out = capsys.readouterr().out
    # Summary still shows truncation hint.
    assert "use -v to see all" in out
    # Verbose block continues from #6.
    assert "Full warning list (8 total)" in out
    # Numbered lines for items 6..8 of the input list (warn-5..warn-7).
    assert "  6. warn-5" in out
    assert "  7. warn-6" in out
    assert "  8. warn-7" in out
    # Top 5 are NOT renumbered into the verbose block.
    assert "  1. warn-0" not in out


def test_verbose_skipped_when_few_warnings(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """If warnings ≤ SUMMARY_WARN_LIMIT, verbose block is not printed."""
    fake_xlsx = _existing_xlsx(tmp_path)
    pipeline = _fake_pipeline(warnings=[f"warn-{i}" for i in range(SUMMARY_WARN_LIMIT)])
    with patch("etl.cli.writeback.run_full_pipeline", return_value=pipeline):
        main([str(fake_xlsx), "-v"])

    out = capsys.readouterr().out
    assert "Full warning list" not in out


def test_summary_includes_reference_count(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Summary block reports reference count + sorted unique categories."""
    fake_xlsx = _existing_xlsx(tmp_path)
    pipeline = _fake_pipeline()

    def _ref(category: str) -> ReferenceText:
        return ReferenceText(
            category=category,
            year=2025,
            month=3,
            content=f"some content for {category}",
            source_file="src.xlsx",
            source_row=92,
            vintage_date=datetime(2026, 1, 31),
            report_type="operational",
            source_priority=10,
        )

    pipeline.canonical_reference = [
        _ref("subsistence_minimum"),
        _ref("min_wage"),
        _ref("fuel_diesel"),
    ]

    with patch("etl.cli.writeback.run_full_pipeline", return_value=pipeline):
        rc = main([str(fake_xlsx)])

    assert rc == 0
    out = capsys.readouterr().out
    assert "Reference texts:  3" in out
    # Categories appear sorted (alphabetical) in the summary line.
    assert "fuel_diesel" in out
    assert "min_wage" in out
    assert "subsistence_minimum" in out
