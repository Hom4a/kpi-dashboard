"""Contract tests for ``extract_salary_block_osnovni``.

Each test builds a synthetic worksheet through ``openpyxl.Workbook`` —
no disk fixtures, deterministic, fast.

Geometry mirrors production «Основні показники» format:

  row 1: ignored (banner)
  row 3: salary header — col A = "Середня з/п по філіях…",
                          col B..F = "2022 рік"…"2026 рік",
                          col G    = "<місяць> <рік>"  (current month label),
                          col H    = "Середня з/п в регіоні (Мінфін)"
  row 4+: branch data rows OR sentinels (footnote / blank / Довідково)
"""
from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from etl.models import SalaryValue
from etl.parsers_salary import extract_salary_block_osnovni
from etl.report_metadata import ReportMetadata

ANNUAL_COLUMNS: tuple[tuple[int, int, bool], ...] = (
    (2, 2022, False),
    (3, 2023, False),
    (4, 2024, False),
    (5, 2025, False),
    (6, 2026, True),
)


def _meta(
    *,
    report_type: str = "operational",
    priority: int = 10,
    vintage: datetime | None = None,
) -> ReportMetadata:
    return ReportMetadata(
        report_type=report_type,  # type: ignore[arg-type]
        vintage_date=vintage or datetime(2026, 4, 10),
        source_priority=priority,
    )


def _build_osnovni_ws(
    branch_rows: list[dict[str, Any]],
    *,
    current_month_label: str | None = "березень 2026",
    extra_rows: list[dict[str, Any]] | None = None,
) -> tuple[Worksheet, int]:
    """Build a synthetic osnovni salary sheet.

    branch_rows: dicts with ``name`` (col A), ``annuals`` (5-list for
                 cols B..F = 2022..2026), ``monthly`` (col G value),
                 ``region`` (col H value). Any field can be omitted.
    extra_rows:  injected after branch_rows; dict ``a`` for col A,
                 ``b`` for col B.

    Returns: (ws, salary_header_row).
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Основні показники"

    salary_header_row = 3
    ws.cell(salary_header_row, 1).value = (
        "Середня з/п по філіях одного штатного працівника, грн"
    )
    for col, year in enumerate(range(2022, 2027), start=2):
        ws.cell(salary_header_row, col).value = f"{year} рік"
    if current_month_label is not None:
        ws.cell(salary_header_row, 7).value = current_month_label
    ws.cell(salary_header_row, 8).value = (
        "Середня з/п в регіоні (за січень 2026, дані Мінфіну)"
    )

    next_row = salary_header_row + 1
    for br in branch_rows:
        ws.cell(next_row, 1).value = br.get("name")
        annuals = br.get("annuals", [None] * 5)
        for offset, val in enumerate(annuals):
            ws.cell(next_row, 2 + offset).value = val
        if "monthly" in br:
            ws.cell(next_row, 7).value = br["monthly"]
        if "region" in br:
            ws.cell(next_row, 8).value = br["region"]
        next_row += 1

    for extra in extra_rows or []:
        if "a" in extra:
            ws.cell(next_row, 1).value = extra["a"]
        if "b" in extra:
            ws.cell(next_row, 2).value = extra["b"]
        next_row += 1

    return ws, salary_header_row


def _extract(
    ws: Worksheet,
    header: int,
    *,
    current_year: int | None = 2026,
    current_month: int | None = 3,
) -> tuple[list[SalaryValue], list[str]]:
    return extract_salary_block_osnovni(
        ws,
        annual_columns=ANNUAL_COLUMNS,
        current_month_col=7,
        current_year=current_year,
        current_month=current_month,
        source_file="synthetic.xlsx",
        base_meta=_meta(),
        start_row=header,
    )


# ---------------------------------------------------------------------------
# 1. Happy path — full row
# ---------------------------------------------------------------------------

def test_happy_path_one_branch_all_years_plus_month() -> None:
    """One branch with values in every column → 5 annual + 1 monthly = 6 emits."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "Карпатський лісовий офіс",
                "annuals": [18874.0, 18819.0, 21646.0, 27796.0, 26125.0],
                "monthly": 39600.0,
                "region": 23963.0,
            },
        ],
    )
    rows, warnings = _extract(ws, hdr)

    assert len(rows) == 6
    assert warnings == []
    annuals = [s for s in rows if s.month == 0]
    monthly = [s for s in rows if s.month == 3]
    assert sorted(s.year for s in annuals) == [2022, 2023, 2024, 2025, 2026]
    assert len(monthly) == 1
    assert monthly[0].year == 2026
    assert monthly[0].salary_uah == 39600.0


# ---------------------------------------------------------------------------
# 2. Sparse old branch — only some annual columns populated
# ---------------------------------------------------------------------------

def test_branch_with_only_recent_years_skips_old() -> None:
    """C2,C3 empty; C4,C5,C6 numeric; C7 numeric → 4 emits."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "філія \"Лісові репродуктивні ресурси\"",
                "annuals": [None, None, 24388.0, 34002.0, 37103.0],
                "monthly": 53338.0,
            },
        ],
    )
    rows, _ = _extract(ws, hdr)

    assert len(rows) == 4
    years_seen = sorted({(s.year, s.month) for s in rows})
    assert years_seen == [(2024, 0), (2025, 0), (2026, 0), (2026, 3)]


# ---------------------------------------------------------------------------
# 3. 2026 YTD column → month=0 (Q2.B confirmation)
# ---------------------------------------------------------------------------

def test_2026_ytd_emitted_as_month_zero() -> None:
    """C6 ('2026 рік' YTD) emits as ``month=0`` regardless of partial-year status."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "Карпатський",
                "annuals": [None, None, None, None, 26125.0],  # only 2026 YTD
                # no monthly to keep the assertion focused
            },
        ],
    )
    rows, _ = _extract(ws, hdr, current_year=None, current_month=None)

    assert len(rows) == 1
    assert rows[0].year == 2026
    assert rows[0].month == 0
    assert rows[0].salary_uah == 26125.0


# ---------------------------------------------------------------------------
# 4. Current month emit
# ---------------------------------------------------------------------------

def test_current_month_emit_uses_parsed_month() -> None:
    """C7 + ``current=(3, 2026)`` → ``SalaryValue(year=2026, month=3)``."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "Подільський",
                "annuals": [None] * 5,
                "monthly": 39600.0,
            },
        ],
    )
    rows, _ = _extract(ws, hdr, current_year=2026, current_month=3)

    assert len(rows) == 1
    assert rows[0].year == 2026
    assert rows[0].month == 3
    assert rows[0].salary_uah == 39600.0


# ---------------------------------------------------------------------------
# 5. current=(None, None) → monthly emit suppressed
# ---------------------------------------------------------------------------

def test_current_month_skipped_when_header_unparseable() -> None:
    """If ``_parse_current_month`` failed, C7 is silently ignored."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "Карпатський",
                "annuals": [18874.0, 18819.0, 21646.0, 27796.0, 26125.0],
                "monthly": 39600.0,  # would have emitted with valid current_*
            },
        ],
    )
    rows, _ = _extract(ws, hdr, current_year=None, current_month=None)

    assert len(rows) == 5
    assert all(s.month == 0 for s in rows)


# ---------------------------------------------------------------------------
# 6. Region always None (Q1.D — Minfin lag not parsed)
# ---------------------------------------------------------------------------

def test_region_always_none_in_osnovni() -> None:
    """Even with a numeric C8, ``region_avg_uah`` stays None for all emits."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "Карпатський",
                "annuals": [18874.0, 18819.0, 21646.0, 27796.0, 26125.0],
                "monthly": 39600.0,
                "region": 23963.0,  # populated, but design says ignore
            },
        ],
    )
    rows, _ = _extract(ws, hdr)

    assert len(rows) == 6
    assert all(s.region_avg_uah is None for s in rows)


# ---------------------------------------------------------------------------
# 7. "-" in region cell does not break extraction
# ---------------------------------------------------------------------------

def test_dash_in_region_does_not_break_extraction() -> None:
    """Production row 80 has C8='-'. Extractor must not crash and must
    still emit the branch's salary cells (region path is read-only and
    ultimately discarded per Q1.D)."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "філія \"Лісові репродуктивні ресурси\"",
                "annuals": [None, None, 24388.0, 34002.0, 37103.0],
                "monthly": 53338.0,
                "region": "-",
            },
        ],
    )
    rows, _ = _extract(ws, hdr)

    assert len(rows) == 4
    assert all(s.region_avg_uah is None for s in rows)


# ---------------------------------------------------------------------------
# 8. Closed branch row skipped entirely
# ---------------------------------------------------------------------------

def test_closed_branch_skipped_entirely() -> None:
    """Every numeric cell is ``"всі філії закриті"`` → 0 emits, 0 warnings."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "філія \"Закрита\"",
                "annuals": ["всі філії закриті"] * 5,
                "monthly": "всі філії закриті",
            },
        ],
    )
    rows, warnings = _extract(ws, hdr)

    assert rows == []
    assert warnings == []


# ---------------------------------------------------------------------------
# 9. Branch name with trailing ** (production row 79)
# ---------------------------------------------------------------------------

def test_branch_name_with_double_star_preserved() -> None:
    """``'філія "..."**'`` is verbatim preserved AND does NOT trigger
    footnote stop (which checks ``startswith('*')``, not ``contains``)."""
    raw = 'філія "Лісовий навчальний центр"**'
    ws, hdr = _build_osnovni_ws(
        [{"name": raw, "annuals": [None, None, 24770.0, 25599.0, 34831.0],
          "monthly": 53483.0}],
    )
    rows, _ = _extract(ws, hdr)

    assert rows
    assert all(s.branch_name == raw for s in rows)


# ---------------------------------------------------------------------------
# 10. Footnote line stops extraction
# ---------------------------------------------------------------------------

def test_footnote_row_stops_extraction() -> None:
    """``*дані відсутні`` row breaks the loop; later branches are NOT seen."""
    ws, hdr = _build_osnovni_ws(
        [
            {
                "name": "Активна",
                "annuals": [None, None, 24770.0, 25599.0, 34831.0],
                "monthly": 53483.0,
            },
        ],
        extra_rows=[
            {"a": "*дані відсутні"},
            {"a": "Привид"},  # would emit if not stopped
        ],
    )
    # Inject ghost monthly to prove break
    ghost_row = hdr + 1 + 1 + 1 + 1  # header+branch+footnote+ghost
    for col in (2, 3, 4, 5, 6, 7):
        ws.cell(ghost_row, col).value = 99999

    rows, _ = _extract(ws, hdr)

    assert {s.branch_name for s in rows} == {"Активна"}
