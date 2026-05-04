"""Contract tests for ``extract_salary_block``.

Each test builds a synthetic worksheet in-memory through ``openpyxl.Workbook``
so the module's behaviour is exercised without disk fixtures: tests stay
fast, deterministic, and immune to source-file regressions.

Geometry of the synthetic salary sheet (mirrors production yearly file):

  row 1: ignored (banner)
  row 2: ignored
  row 3: salary header — col A = "Середня з/п по лісових офісах…",
                          col B..M = datetime per month,
                          col N    = "Середня за рік",
                          col O    = "Середня з/п в регіоні"
  row 4+: branch data rows OR sentinels (footnote / blank / Довідково)
"""
from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from etl.models import SalaryValue
from etl.parsers_salary import extract_salary_block
from etl.report_metadata import ReportMetadata


def _meta(
    *,
    report_type: str = "operational",
    priority: int = 10,
    vintage: datetime | None = None,
) -> ReportMetadata:
    """ReportMetadata constructor with sensible defaults."""
    return ReportMetadata(
        report_type=report_type,  # type: ignore[arg-type]
        vintage_date=vintage or datetime(2026, 1, 31),
        source_priority=priority,
    )


def _build_ws(
    branch_rows: list[dict[str, Any]],
    *,
    year: int = 2025,
    has_region_header: bool = True,
    extra_rows: list[dict[str, Any]] | None = None,
) -> tuple[Worksheet, dict[int, tuple[int, int]], int]:
    """Build a synthetic salary sheet.

    branch_rows: list of dicts with optional keys ``name`` (col A),
                 ``monthly`` (12-list of cell values for B..M),
                 ``annual`` (col N value),
                 ``region`` (col O value).
    extra_rows:  list of dicts with key ``a`` for col A, optional ``b``
                 for col B — used to inject sentinels (footnotes,
                 blanks, дoвідково) AFTER the branch rows.

    Returns: (ws, month_map, salary_header_row).
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = str(year)

    # Salary header at row 3.
    salary_header_row = 3
    ws.cell(salary_header_row, 1).value = (
        "Середня з/п по лісових офісах одного штатного працівника, грн"
    )
    month_map: dict[int, tuple[int, int]] = {}
    for col, month in enumerate(range(1, 13), start=2):
        ws.cell(salary_header_row, col).value = datetime(year, month, 28)
        month_map[col] = (year, month)
    ws.cell(salary_header_row, 14).value = "Середня за рік"
    if has_region_header:
        ws.cell(salary_header_row, 15).value = "Середня з/п в регіоні"

    next_row = salary_header_row + 1
    for br in branch_rows:
        ws.cell(next_row, 1).value = br.get("name")
        monthly = br.get("monthly", [None] * 12)
        for offset, val in enumerate(monthly):
            ws.cell(next_row, 2 + offset).value = val
        if "annual" in br:
            ws.cell(next_row, 14).value = br["annual"]
        if "region" in br:
            ws.cell(next_row, 15).value = br["region"]
        next_row += 1

    for extra in extra_rows or []:
        if "a" in extra:
            ws.cell(next_row, 1).value = extra["a"]
        if "b" in extra:
            ws.cell(next_row, 2).value = extra["b"]
        next_row += 1

    return ws, month_map, salary_header_row


def _full_year(start: float = 25000.0) -> list[float]:
    """12 monthly salary values starting at ``start`` and incrementing."""
    return [start + i * 100 for i in range(12)]


def _extract(
    ws: Worksheet, month_map: dict[int, tuple[int, int]], header: int
) -> tuple[list[SalaryValue], list[str]]:
    """Convenience invoker with default metadata."""
    return extract_salary_block(
        ws,
        month_map=month_map,
        ytd_year=2025,
        source_file="synthetic.xlsx",
        base_meta=_meta(report_type="operational", priority=10),
        ytd_meta=_meta(report_type="accounting_ytd", priority=20),
        start_row=header,
    )


# ---------------------------------------------------------------------------
# 1. Happy path
# ---------------------------------------------------------------------------

def test_happy_path_one_branch_full_year() -> None:
    """One active branch with all 12 monthly + annual + region → 13 emits."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Карпатський лісовий офіс",
                "monthly": _full_year(),
                "annual": 26500.0,
                "region": 22000.0,
            },
        ],
    )
    rows, warnings = _extract(ws, mm, hdr)

    assert len(rows) == 13
    assert warnings == []
    months_seen = sorted(s.month for s in rows)
    assert months_seen == [0] + list(range(1, 13))
    assert all(s.branch_name == "Карпатський лісовий офіс" for s in rows)
    assert all(s.region_avg_uah == 22000.0 for s in rows)
    annual = next(s for s in rows if s.month == 0)
    assert annual.salary_uah == 26500.0


# ---------------------------------------------------------------------------
# 2. Closed branch row skipped entirely (variant B)
# ---------------------------------------------------------------------------

def test_closed_branch_row_skipped_entirely() -> None:
    """A row whose every numeric cell is ``"всі філії закриті"`` emits nothing."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "філія \"Закрита\"",
                "monthly": ["всі філії закриті"] * 12,
                "annual": "всі філії закриті",
                "region": None,
            },
        ],
    )
    rows, warnings = _extract(ws, mm, hdr)

    assert rows == []
    assert warnings == []  # variant B — no warning for closed branch


# ---------------------------------------------------------------------------
# 3. Mixed active + closed branches
# ---------------------------------------------------------------------------

def test_mixed_active_and_closed_branches() -> None:
    """Active branch yields 13 emits; closed branch yields 0; both coexist."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Активна",
                "monthly": _full_year(),
                "annual": 26500.0,
                "region": 22000.0,
            },
            {
                "name": "Закрита",
                "monthly": ["всі філії закриті"] * 12,
                "annual": "всі філії закриті",
            },
        ],
    )
    rows, warnings = _extract(ws, mm, hdr)

    assert len(rows) == 13
    assert {s.branch_name for s in rows} == {"Активна"}
    assert warnings == []


# ---------------------------------------------------------------------------
# 4. Missing region column (region cell unpopulated)
# ---------------------------------------------------------------------------

def test_missing_region_column_yields_none() -> None:
    """Branch without a region cell → ``region_avg_uah=None`` for all 13 emits."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Полісся",
                "monthly": _full_year(20000.0),
                "annual": 21000.0,
                # no "region" key — col O empty
            },
        ],
    )
    rows, _ = _extract(ws, mm, hdr)

    assert len(rows) == 13
    assert all(s.region_avg_uah is None for s in rows)


# ---------------------------------------------------------------------------
# 5. Empty cell mid-year — only that period skipped
# ---------------------------------------------------------------------------

def test_empty_cell_mid_year_skipped() -> None:
    """One ``None`` mid-year drops one monthly emit; rest of branch unaffected."""
    base = _full_year()
    monthly: list[float | None] = list(base)
    monthly[5] = None  # June (col G) blank
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Подільський",
                "monthly": monthly,
                "annual": 27000.0,
                "region": 22500.0,
            },
        ],
    )
    rows, _ = _extract(ws, mm, hdr)

    # 11 monthly + 1 annual = 12
    assert len(rows) == 12
    months = sorted(s.month for s in rows)
    assert months == [0, 1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12]


# ---------------------------------------------------------------------------
# 6. Two consecutive empty rows terminate the section
# ---------------------------------------------------------------------------

def test_two_consecutive_empty_rows_stop() -> None:
    """Two blank rows in a row → break; later branches are NOT seen."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Перша",
                "monthly": _full_year(20000.0),
                "annual": 21000.0,
            },
        ],
        extra_rows=[
            {},   # row 5 blank
            {},   # row 6 blank → break
            # The branch below row 6 must be ignored.
        ],
    )
    # Manually inject a post-break branch by writing directly.
    last_blank_row = hdr + 1 + 2  # branch=4, blanks=5,6 → ghost branch=7
    ws.cell(last_blank_row + 1, 1).value = "Привид"
    for offset in range(12):
        ws.cell(last_blank_row + 1, 2 + offset).value = 99999

    rows, _ = _extract(ws, mm, hdr)

    branch_names = {s.branch_name for s in rows}
    assert "Перша" in branch_names
    assert "Привид" not in branch_names


# ---------------------------------------------------------------------------
# 7. Footnote (`*тощо`) terminates extraction
# ---------------------------------------------------------------------------

def test_footnote_star_stops_extraction() -> None:
    """Row starting with ``*`` ends the section — no later rows scanned."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Карпатський",
                "monthly": _full_year(),
                "annual": 26500.0,
            },
        ],
        extra_rows=[
            {"a": "*тощо — пояснення"},
            {"a": "Привид після футера"},  # would emit if not stopped
        ],
    )
    # Add fake monthly to ghost so absence proves the break.
    ghost_row = hdr + 1 + 1 + 1 + 1  # header+branch+footnote+ghost
    for offset in range(12):
        ws.cell(ghost_row, 2 + offset).value = 88888

    rows, _ = _extract(ws, mm, hdr)

    assert {s.branch_name for s in rows} == {"Карпатський"}


# ---------------------------------------------------------------------------
# 8. Defensive: «Довідково:» header terminates extraction
# ---------------------------------------------------------------------------

def test_dovidkovo_header_stops_extraction() -> None:
    """Stray reference header inside salary section → break (defensive)."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Активна",
                "monthly": _full_year(),
                "annual": 26500.0,
            },
        ],
        extra_rows=[
            {"a": "Довідково:"},
            {"a": "Не повинно потрапити"},
        ],
    )
    rows, _ = _extract(ws, mm, hdr)

    assert {s.branch_name for s in rows} == {"Активна"}


# ---------------------------------------------------------------------------
# 9. No salary header → warning, no emits
# ---------------------------------------------------------------------------

def test_no_salary_header_returns_warning() -> None:
    """Worksheet without keyword → ``no_salary_block_found`` info warning."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(1, 1).value = "Some other content"
    ws.cell(2, 1).value = "Загальна реалізація"

    rows, warnings = extract_salary_block(
        ws,
        month_map={2 + i: (2025, i + 1) for i in range(12)},
        ytd_year=2025,
        source_file="synthetic.xlsx",
        base_meta=_meta(),
        ytd_meta=_meta(report_type="accounting_ytd", priority=20),
        start_row=None,  # let the function scan and miss
    )

    assert rows == []
    assert warnings == ["no_salary_block_found"]


# ---------------------------------------------------------------------------
# 10. Branch name verbatim (quotes preserved)
# ---------------------------------------------------------------------------

def test_branch_name_verbatim_with_quotes() -> None:
    """Quoted branch name kept exactly — repository handles normalisation."""
    raw = 'Філія "Карпатський лісовий офіс"'
    ws, mm, hdr = _build_ws(
        [{"name": raw, "monthly": _full_year(), "annual": 26500.0}],
    )
    rows, _ = _extract(ws, mm, hdr)

    assert rows
    assert all(s.branch_name == raw for s in rows)


# ---------------------------------------------------------------------------
# 11. month=0 row uses ytd_meta; monthly rows use base_meta
# ---------------------------------------------------------------------------

def test_annual_avg_uses_ytd_meta() -> None:
    """``month=0`` emit takes ``accounting_ytd``/``priority=20``;
    months 1..12 take ``operational``/``priority=10``."""
    ws, mm, hdr = _build_ws(
        [{"name": "Карпатський", "monthly": _full_year(), "annual": 26500.0}],
    )
    rows, _ = _extract(ws, mm, hdr)

    annual = [s for s in rows if s.month == 0]
    monthly = [s for s in rows if s.month != 0]

    assert len(annual) == 1
    assert annual[0].report_type == "accounting_ytd"
    assert annual[0].source_priority == 20

    assert len(monthly) == 12
    assert all(s.report_type == "operational" for s in monthly)
    assert all(s.source_priority == 10 for s in monthly)


# ---------------------------------------------------------------------------
# 12. Region propagated to all 13 emits of the same branch
# ---------------------------------------------------------------------------

def test_region_propagated_to_all_rows_of_branch() -> None:
    """Single region cell → identical ``region_avg_uah`` on all 13 emits."""
    ws, mm, hdr = _build_ws(
        [
            {
                "name": "Карпатський",
                "monthly": _full_year(),
                "annual": 26500.0,
                "region": 22000.0,
            },
        ],
    )
    rows, _ = _extract(ws, mm, hdr)

    assert len(rows) == 13
    assert all(s.region_avg_uah == 22000.0 for s in rows)
