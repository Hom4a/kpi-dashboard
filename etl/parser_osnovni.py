"""Parser for «Основні показники діяльності» — format B (wide by year).

Sheet layout:
  - one worksheet (first sheet)
  - header row: B="2022 рік", C="2023 рік", D="2024 рік", E="2025 рік",
                F="2026 рік" (YTD), G="<місяць> 2026" (current month)
  - below header: rows with indicator name in column A, values in B..G
  - multiple logical sections separated by blank rows / section titles

All emitted fact rows carry revision metadata from ``infer_report_metadata``.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .metrics import is_ignored, resolve_metric, resolve_species
from .models import (
    AnimalValue,
    AnnualValue,
    MonthlyValue,
    ParseResult,
    ReferenceText,
    SalaryValue,
    SpeciesAnnual,
    SpeciesMonthly,
)
from .parsers_animals import (
    _ANIMALS_HEADER_KEYWORD,
    extract_animals_block_osnovni,
)
from .parsers_reference import extract_reference_block
from .parsers_salary import extract_salary_block_osnovni
from .report_metadata import infer_report_metadata
from .utils import parse_composite_cell, safe_number

MONTHS_UA: dict[str, int] = {
    "січень": 1, "лютий": 2, "березень": 3, "квітень": 4, "травень": 5,
    "червень": 6, "липень": 7, "серпень": 8, "вересень": 9,
    "жовтень": 10, "листопад": 11, "грудень": 12,
}

# Column map for header row: (col_idx, year, is_ytd)
ANNUAL_COLUMNS: tuple[tuple[int, int, bool], ...] = (
    (2, 2022, False),
    (3, 2023, False),
    (4, 2024, False),
    (5, 2025, False),
    (6, 2026, True),
)

_SECTION_HEADERS: frozenset[str] = frozenset(
    {
        "показники",
        "в тому числі:",
        "довідково",
        "довідково:",
    }
)

# Mirrors parsers_salary._SALARY_HEADER_KEYWORD; kept duplicate (not imported)
# because parser_osnovni decides whether to *capture* the row index or *skip*
# the section, while parsers_salary decides whether to *find* the header.
_SALARY_SECTION_MARKER = "середня з/п по"

# Same shape as the salary marker — animals section header is
# 'Чисельність/кількість лімітів'. Re-exported from parsers_animals
# under a parser-local name for readability.
_ANIMALS_SECTION_MARKER = _ANIMALS_HEADER_KEYWORD


def _find_header_row(ws: object, max_scan: int = 10) -> int | None:
    for row_idx in range(1, max_scan + 1):
        b_val = ws.cell(row_idx, 2).value  # type: ignore[attr-defined]
        if b_val is None:
            continue
        b_str = str(b_val).strip().lower()
        if "2022" in b_str and ("рік" in b_str or "р." in b_str):
            return row_idx
    return None


def _parse_current_month(header_value: str) -> tuple[int | None, int | None]:
    low = header_value.strip().lower()
    month = next((num for name, num in MONTHS_UA.items() if name in low), None)
    y_match = re.search(r"(20\d{2})", low)
    year = int(y_match.group(1)) if y_match else None
    return month, year


def parse_osnovni_annual(path: str | Path) -> ParseResult:
    """Parse an Excel «Основні показники» file (format B, wide by year)."""
    path_str = str(path)
    meta = infer_report_metadata(path_str)
    wb = load_workbook(path_str, data_only=True)
    ws = wb.worksheets[0]

    header_row = _find_header_row(ws)
    if header_row is None:
        return ParseResult(errors=["header_row_not_found"])

    current_month_header = str(ws.cell(header_row, 7).value or "").strip()
    current_month, current_year = _parse_current_month(current_month_header)

    annual: list[AnnualValue] = []
    monthly: list[MonthlyValue] = []
    species_annual: list[SpeciesAnnual] = []
    species_monthly: list[SpeciesMonthly] = []
    reference: list[ReferenceText] = []
    salary: list[SalaryValue] = []
    salary_header_row: int | None = None
    animal: list[AnimalValue] = []
    animals_header_row: int | None = None
    warnings: list[str] = list(meta.warnings)
    errors: list[str] = []

    if current_month is None or current_year is None:
        warnings.append(
            f"could not parse current month header: '{current_month_header}'"
        )

    # Pick coordinates for reference rows: monthly snapshot if the header
    # carried a current-month label, otherwise fall back to the last YTD
    # year with month=0 (annual snapshot).
    if current_year is not None and current_month is not None:
        ref_year, ref_month = current_year, current_month
    else:
        ref_year, ref_month = ANNUAL_COLUMNS[-1][1], 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        a_raw = ws.cell(row_idx, 1).value
        if a_raw is None or str(a_raw).strip() == "":
            continue

        a_name = str(a_raw).strip()
        a_norm = " ".join(a_name.split()).casefold()

        # Salary section is parsed separately by extract_salary_block_osnovni;
        # capturing the header row here lets us hand it off without rescanning.
        # Must run before _SECTION_HEADERS check (the salary header isn't a
        # section banner but is still distinct from metric rows).
        if _SALARY_SECTION_MARKER in a_norm:
            salary_header_row = row_idx
            break

        if a_norm in _SECTION_HEADERS:
            continue
        if a_norm.startswith(_ANIMALS_SECTION_MARKER):
            # Capture for post-loop extract_animals_block_osnovni. Don't
            # break — salary section still lies further down (row 69).
            if animals_header_row is None:
                animals_header_row = row_idx
            continue
        if a_norm.startswith("чисельність/кількість"):
            # Defensive fallback for any other 'чисельність/кількість*'
            # header that may not include the 'лімітів' suffix.
            continue
        if a_norm.startswith("*"):
            continue
        if is_ignored(a_name):
            # Derived metric — computed post-canonical, not read from Excel.
            continue

        species_code = resolve_species(a_name)
        if species_code is not None:
            for col_idx, year, _is_ytd in ANNUAL_COLUMNS:
                vol, price, warn = parse_composite_cell(ws.cell(row_idx, col_idx).value)
                if vol is not None or price is not None:
                    species_annual.append(
                        SpeciesAnnual(
                            species=species_code,
                            year=year,
                            volume_km3=vol,
                            avg_price_grn=price,
                            source_file=path_str,
                            source_row=row_idx,
                            vintage_date=meta.vintage_date,
                            report_type=meta.report_type,
                            source_priority=meta.source_priority,
                        )
                    )
                if warn and warn not in ("empty_marker", "single_value"):
                    warnings.append(
                        f"row {row_idx} col {col_idx} species {species_code}: {warn}"
                    )
            if current_month is not None and current_year is not None:
                vol, price, warn = parse_composite_cell(ws.cell(row_idx, 7).value)
                if vol is not None or price is not None:
                    species_monthly.append(
                        SpeciesMonthly(
                            species=species_code,
                            year=current_year,
                            month=current_month,
                            volume_km3=vol,
                            avg_price_grn=price,
                            source_file=path_str,
                            source_row=row_idx,
                            vintage_date=meta.vintage_date,
                            report_type=meta.report_type,
                            source_priority=meta.source_priority,
                        )
                    )
                if warn and warn not in ("empty_marker", "single_value"):
                    warnings.append(
                        f"row {row_idx} col 7 species {species_code}: {warn}"
                    )
            continue

        metric_code = resolve_metric(a_name)
        if metric_code is None:
            warnings.append(f"row {row_idx}: unknown_metric '{a_name}'")
            continue

        for col_idx, year, is_ytd in ANNUAL_COLUMNS:
            val, warn, raw = safe_number(ws.cell(row_idx, col_idx).value)
            if val is not None or raw is not None:
                annual.append(
                    AnnualValue(
                        metric_code=metric_code,
                        year=year,
                        value=val,
                        value_text=raw,
                        is_ytd=is_ytd,
                        source_file=path_str,
                        source_row=row_idx,
                        vintage_date=meta.vintage_date,
                        report_type=meta.report_type,
                        source_priority=meta.source_priority,
                    )
                )
            if warn and warn != "empty_marker":
                warnings.append(f"row {row_idx} col {col_idx} {metric_code}: {warn}")

        if current_month is not None and current_year is not None:
            val, warn, raw = safe_number(ws.cell(row_idx, 7).value)
            if val is not None or raw is not None:
                monthly.append(
                    MonthlyValue(
                        metric_code=metric_code,
                        year=current_year,
                        month=current_month,
                        value=val,
                        value_text=raw,
                        source_file=path_str,
                        source_row=row_idx,
                        vintage_date=meta.vintage_date,
                        report_type=meta.report_type,
                        source_priority=meta.source_priority,
                    )
                )
            if warn and warn != "empty_marker":
                warnings.append(f"row {row_idx} col 7 {metric_code}: {warn}")

    # «Середня з/п по філіях» — main loop captured the header row before
    # breaking; extractor walks downward from there. C2..C6 → month=0 per
    # year, C7 → (current_year, current_month). C8 (Minfin region with
    # 2-month lag) is intentionally not parsed — see Q1.D in 5.4.3.
    sal_rows, sal_warns = extract_salary_block_osnovni(
        ws,
        annual_columns=ANNUAL_COLUMNS,
        current_month_col=7,
        current_year=current_year,
        current_month=current_month,
        source_file=path_str,
        base_meta=meta,
        start_row=salary_header_row,
    )
    salary.extend(sal_rows)
    warnings.extend(sal_warns)

    # «Чисельність/кількість лімітів» — animals census, captured by main
    # loop. Years come from ANNUAL_COLUMNS; the is_ytd column (2026) is
    # filtered inside the extractor — animals reflect closed years only.
    if animals_header_row is not None:
        anim_rows, anim_warns = extract_animals_block_osnovni(
            ws,
            annual_columns=ANNUAL_COLUMNS,
            source_file=path_str,
            base_meta=meta,
            header_row=animals_header_row,
        )
        animal.extend(anim_rows)
        warnings.extend(anim_warns)

    # «Довідково» — extracted post-loop, independent of metric scanning.
    # ws here is worksheets[0] only (osnovni multi-sheet copies on
    # year-sheets are not iterated, satisfying edge-case E1).
    ref_rows, ref_warns = extract_reference_block(
        ws,
        year=ref_year,
        month=ref_month,
        source_file=path_str,
        vintage_date=meta.vintage_date,
        report_type=meta.report_type,
        source_priority=meta.source_priority,
    )
    reference.extend(ref_rows)
    warnings.extend(ref_warns)

    return ParseResult(
        annual=annual,
        monthly=monthly,
        species_annual=species_annual,
        species_monthly=species_monthly,
        reference=reference,
        salary=salary,
        animal=animal,
        warnings=warnings,
        errors=errors,
    )
