"""Parser for «<Year> рік.xlsx» — format A (wide by month).

Sheet layout:
  - one worksheet (first sheet), usually named "2023"/"2024"/"2025"
  - header row: B..M contain 12 monthly dates (Excel serials OR datetime),
                N = "Від початку року" (YTD), O = "Формули" (literal text)
  - below header: indicator name in col A, monthly values in B..M, YTD in N
  - multiple sections separated by blank rows: operational → taxes → salary
  - salary-by-branch section is intentionally skipped (TODO: separate model)

Revision metadata:
  - B..M monthly cells inherit the file's base metadata (operational/10)
  - N (YTD) is bumped to ``accounting_ytd`` / priority=20 via ``ytd_override``
    so canonical view prefers the accountant-authored total over the
    month-by-month sum (reversals, reclassifications).
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .metrics import is_ignored, resolve_metric, resolve_species
from .models import (
    AnnualValue,
    MonthlyValue,
    ParseResult,
    ReferenceText,
    SalaryValue,
    SpeciesAnnual,
    SpeciesMonthly,
)
from .parsers_reference import extract_reference_block
from .parsers_salary import extract_salary_block
from .report_metadata import ReportMetadata, infer_report_metadata
from .utils import parse_composite_cell, safe_number

MONTH_COL_START, MONTH_COL_END = 2, 13  # columns B..M (12 months)
YTD_COL = 14  # column N
FORMULA_COL = 15  # column O — ignored

_SECTION_HEADERS: frozenset[str] = frozenset(
    {
        "показники",
        "в тому числі:",
        "довідково",
        "довідково:",
    }
)

# TODO: branch salaries need a separate BranchSalary model + dim_branch
# dictionary; handled in a later step. For now we stop scanning here.
_SALARY_SECTION_MARKER = "середня з/п по лісових офісах"

_REFERENCE_PREFIXES: tuple[str, ...] = (
    "прожитковий мінімум",
    "мінімальна заробітна",
    "середня заробітна плата країна",
)


def _to_date(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and 40000 <= float(value) <= 60000:
        return from_excel(float(value))
    return None


def _find_header_row(ws: object, max_scan: int = 10) -> int | None:
    for row_idx in range(1, max_scan + 1):
        b_val = ws.cell(row_idx, MONTH_COL_START).value  # type: ignore[attr-defined]
        if _to_date(b_val) is not None:
            return row_idx
    return None


def _build_month_map(ws: object, header_row: int) -> tuple[dict[int, tuple[int, int]], str | None]:
    month_map: dict[int, tuple[int, int]] = {}
    prev_month: int | None = None
    year_seen: int | None = None
    for col in range(MONTH_COL_START, MONTH_COL_END + 1):
        dt = _to_date(ws.cell(header_row, col).value)  # type: ignore[attr-defined]
        if dt is None:
            return {}, f"missing_date_in_header_col_{col}"
        if year_seen is None:
            year_seen = dt.year
        elif dt.year != year_seen:
            return {}, "month_sequence_broken (year mismatch)"
        if prev_month is not None and dt.month != prev_month + 1:
            return {}, f"month_sequence_broken (gap at col {col})"
        month_map[col] = (dt.year, dt.month)
        prev_month = dt.month
    if len(month_map) != 12:
        return {}, "month_sequence_broken (expected 12 months)"
    return month_map, None


def _ytd_meta(base: ReportMetadata) -> ReportMetadata:
    """YTD column uses ``ytd_override`` when provided, else falls back to base."""
    return base.ytd_override if base.ytd_override is not None else base


def parse_annual_monthly(path: str | Path) -> ParseResult:
    """Parse a «<Year> рік.xlsx» file (format A, wide by month)."""
    path_str = str(path)
    base_meta = infer_report_metadata(path_str)
    ytd_meta = _ytd_meta(base_meta)

    wb = load_workbook(path_str, data_only=True)
    ws = wb.worksheets[0]

    header_row = _find_header_row(ws)
    if header_row is None:
        return ParseResult(errors=["header_row_not_found"])

    month_map, err = _build_month_map(ws, header_row)
    if err:
        return ParseResult(errors=[err])

    annual: list[AnnualValue] = []
    monthly: list[MonthlyValue] = []
    species_annual: list[SpeciesAnnual] = []
    species_monthly: list[SpeciesMonthly] = []
    reference: list[ReferenceText] = []
    salary: list[SalaryValue] = []
    salary_header_row: int | None = None
    warnings: list[str] = list(base_meta.warnings)

    sheet_year = next(iter(month_map.values()))[0]

    for row_idx in range(header_row + 1, ws.max_row + 1):
        a_raw = ws.cell(row_idx, 1).value
        if a_raw is None or str(a_raw).strip() == "":
            continue

        a_name = str(a_raw).strip()
        a_norm = " ".join(a_name.split()).casefold()

        if _SALARY_SECTION_MARKER in a_norm:
            salary_header_row = row_idx
            break

        if a_norm in _SECTION_HEADERS:
            continue
        if a_norm.startswith("чисельність/кількість"):
            continue
        if a_norm.startswith("*"):
            continue
        if any(a_norm.startswith(pref) for pref in _REFERENCE_PREFIXES):
            continue
        if is_ignored(a_name):
            # Derived metric — computed post-canonical, not read from Excel.
            continue

        species_code = resolve_species(a_name)
        if species_code is not None:
            for col in range(MONTH_COL_START, MONTH_COL_END + 1):
                year, month = month_map[col]
                vol, price, warn = parse_composite_cell(ws.cell(row_idx, col).value)
                if vol is not None or price is not None:
                    species_monthly.append(
                        SpeciesMonthly(
                            species=species_code,
                            year=year,
                            month=month,
                            volume_km3=vol,
                            avg_price_grn=price,
                            source_file=path_str,
                            source_row=row_idx,
                            vintage_date=base_meta.vintage_date,
                            report_type=base_meta.report_type,
                            source_priority=base_meta.source_priority,
                        )
                    )
                if warn and warn not in ("empty_marker", "single_value"):
                    warnings.append(
                        f"row {row_idx} col {col} species {species_code}: {warn}"
                    )

            vol, price, warn = parse_composite_cell(ws.cell(row_idx, YTD_COL).value)
            if vol is not None or price is not None:
                species_annual.append(
                    SpeciesAnnual(
                        species=species_code,
                        year=sheet_year,
                        volume_km3=vol,
                        avg_price_grn=price,
                        source_file=path_str,
                        source_row=row_idx,
                        vintage_date=ytd_meta.vintage_date,
                        report_type=ytd_meta.report_type,
                        source_priority=ytd_meta.source_priority,
                    )
                )
            if warn and warn not in ("empty_marker", "single_value"):
                warnings.append(
                    f"row {row_idx} col {YTD_COL} species {species_code}: {warn}"
                )
            continue

        metric_code = resolve_metric(a_name)
        if metric_code is None:
            warnings.append(f"row {row_idx}: unknown_metric '{a_name}'")
            continue

        for col in range(MONTH_COL_START, MONTH_COL_END + 1):
            year, month = month_map[col]
            val, warn, raw = safe_number(ws.cell(row_idx, col).value)
            if val is not None or raw is not None:
                monthly.append(
                    MonthlyValue(
                        metric_code=metric_code,
                        year=year,
                        month=month,
                        value=val,
                        value_text=raw,
                        source_file=path_str,
                        source_row=row_idx,
                        vintage_date=base_meta.vintage_date,
                        report_type=base_meta.report_type,
                        source_priority=base_meta.source_priority,
                    )
                )
            if warn and warn != "empty_marker":
                warnings.append(f"row {row_idx} col {col} {metric_code}: {warn}")

        val, warn, raw = safe_number(ws.cell(row_idx, YTD_COL).value)
        if val is not None or raw is not None:
            annual.append(
                AnnualValue(
                    metric_code=metric_code,
                    year=sheet_year,
                    value=val,
                    value_text=raw,
                    is_ytd=False,
                    source_file=path_str,
                    source_row=row_idx,
                    vintage_date=ytd_meta.vintage_date,
                    report_type=ytd_meta.report_type,
                    source_priority=ytd_meta.source_priority,
                )
            )
        if warn and warn != "empty_marker":
            warnings.append(f"row {row_idx} col {YTD_COL} {metric_code}: {warn}")

    # «Середня з/п по лісових офісах» — main loop captured the header row
    # before breaking; extractor walks downward from there.
    sal_rows, sal_warns = extract_salary_block(
        ws,
        month_map=month_map,
        ytd_year=sheet_year,
        source_file=path_str,
        base_meta=base_meta,
        ytd_meta=ytd_meta,
        start_row=salary_header_row,
    )
    salary.extend(sal_rows)
    warnings.extend(sal_warns)

    # «Довідково» — extracted independently of the main metric loop so the
    # SALARY_SECTION_MARKER break above can't mask it. The extractor scans
    # the sheet for its own header (single sheet only — no roaming).
    ref_rows, ref_warns = extract_reference_block(
        ws,
        year=sheet_year,
        month=0,  # annual snapshot — yearly file
        source_file=path_str,
        vintage_date=ytd_meta.vintage_date,
        report_type=ytd_meta.report_type,
        source_priority=ytd_meta.source_priority,
    )
    reference.extend(ref_rows)
    warnings.extend(ref_warns)

    return ParseResult(
        annual=annual,
        monthly=monthly,
        species_annual=species_annual,
        species_monthly=species_monthly,
        reference=reference,
        salary=salary,
        warnings=warnings,
    )


__all__ = ["parse_annual_monthly"]
