"""``python -m etl.cli.writeback`` — push canonical Excel facts to on-prem Postgres.

Default mode is **dry-run**: parses the file, reports what would be
written, and exits without opening any DB connection. ``--commit`` wires
the SSH tunnel and applies the batch transactionally.

Exit codes:
    0  success (dry-run printed plan, or commit succeeded)
    2  parser auto-detect failed
    3  parser produced errors (file structure unexpected)
    4  --commit but ``.env`` is missing or incomplete
    5  DB connection / transport failure
    6  unknown metric_code (no mapping in code_mapper)
"""
from __future__ import annotations

import argparse
import sys
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from etl.canonical import (
    canonical_annual,
    canonical_monthly,
    canonical_reference,
    canonical_species_annual,
    canonical_species_monthly,
)
from etl.derived import compute_derived_annual
from etl.models import (
    AnnualValue,
    MonthlyValue,
    ParseResult,
    ReferenceText,
    SpeciesAnnual,
    SpeciesMonthly,
)
from etl.parser_annual_monthly import parse_annual_monthly
from etl.parser_osnovni import parse_osnovni_annual

if TYPE_CHECKING:
    from etl.db.interface import WriteBatch, WriteResult


# ---------------------------------------------------------------------------
# Parser auto-detection
# ---------------------------------------------------------------------------

ParserName = str  # "annual_monthly" | "osnovni_annual"


_ANNUAL_MONTHLY_DAY_DELTA_MIN = 27  # tightest real month gap (Feb→Mar etc.)
_ANNUAL_MONTHLY_DAY_DELTA_MAX = 32


def _parse_year_cell(value: object) -> int | None:
    """Return the 4-digit year token inside ``value`` if it looks like a
    "20XX рік" header cell, else None.

    Year window is dynamic — anchored on the current year so the heuristic
    doesn't have to be revisited every few years.
    """
    if not isinstance(value, str):
        return None
    low = value.strip().lower()
    if "рік" not in low and "р." not in low:
        return None
    upper = datetime.now().year + 5
    for y in range(2000, upper):
        if str(y) in low:
            return y
    return None


def _parse_date_cell(value: object) -> datetime | None:
    """Return a ``datetime`` if ``value`` is one (or is an Excel serial in
    the plausible 2010-2060 range), else None."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, bool):  # bool is a subclass of int — guard explicitly
        return None
    if isinstance(value, (int, float)) and 40000 <= float(value) <= 60000:
        # Excel serial → datetime via openpyxl helper. Lazy import to
        # keep this CLI testable without round-tripping every cell.
        from openpyxl.utils.datetime import from_excel
        result: datetime = from_excel(float(value))
        return result
    return None


def detect_parser(xlsx_path: Path) -> ParserName:
    """Inspect the workbook header row to choose the right parser.

    Strict criteria (≥2 consecutive cells must agree):

    * ``osnovni_annual`` — cell B and cell C both decode to ``YYYY рік``
      with **strictly different** years (no copy-paste artefacts, no stray
      ``"01.01.2025 рік"`` strings beating the check).
    * ``annual_monthly`` — cell B and cell C both decode to a date
      (datetime or Excel serial) **27-32 days apart** — i.e. consecutive
      months. Two random datetimes side-by-side don't qualify.

    Scans the first ~10 rows because real workbooks sometimes lead with a
    banner row.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    osnovni_reason = "no row had B and C decoding to distinct 'YYYY рік' tokens"
    annual_monthly_reason = (
        "no row had B and C as dates 27-32 days apart"
    )
    try:
        for row_idx in range(1, 11):
            b_val = ws.cell(row_idx, 2).value
            c_val = ws.cell(row_idx, 3).value

            # Osnovni: both B and C are year-headers, with strictly increasing year.
            year_b = _parse_year_cell(b_val)
            year_c = _parse_year_cell(c_val)
            if year_b is not None and year_c is not None and year_c > year_b:
                return "osnovni_annual"

            # Annual-monthly: both B and C decode as dates 27-32 days apart.
            date_b = _parse_date_cell(b_val)
            date_c = _parse_date_cell(c_val)
            if date_b is not None and date_c is not None:
                delta_days = (date_c - date_b).days
                if (
                    _ANNUAL_MONTHLY_DAY_DELTA_MIN
                    <= delta_days
                    <= _ANNUAL_MONTHLY_DAY_DELTA_MAX
                ):
                    return "annual_monthly"
    finally:
        wb.close()
    raise ValueError(
        f"Could not auto-detect parser for {xlsx_path}.\n"
        f"  osnovni_annual rejected: {osnovni_reason}\n"
        f"  annual_monthly rejected: {annual_monthly_reason}"
    )


# ---------------------------------------------------------------------------
# Pipeline: parse → canonical → derived
# ---------------------------------------------------------------------------

@dataclass
class PipelineOutput:
    """Aggregate of everything writeback needs from one file."""

    parser_name: ParserName
    raw_result: ParseResult
    canonical_annual: list[AnnualValue]
    canonical_monthly: list[MonthlyValue]
    canonical_species_annual: list[SpeciesAnnual]
    canonical_species_monthly: list[SpeciesMonthly]
    canonical_reference: list[ReferenceText]
    derived_annual: list[AnnualValue]


def run_full_pipeline(xlsx_path: Path) -> PipelineOutput:
    """Run parser + canonical + derived. Returns aggregated facts."""
    parser_name = detect_parser(xlsx_path)
    parse_fn = (
        parse_osnovni_annual if parser_name == "osnovni_annual" else parse_annual_monthly
    )
    raw = parse_fn(xlsx_path)

    canon_a = canonical_annual(raw.annual)
    canon_m = canonical_monthly(raw.monthly)
    canon_sa = canonical_species_annual(raw.species_annual)
    canon_sm = canonical_species_monthly(raw.species_monthly)
    canon_ref = canonical_reference(raw.reference)
    derived = compute_derived_annual(canon_a)

    return PipelineOutput(
        parser_name=parser_name,
        raw_result=raw,
        canonical_annual=canon_a,
        canonical_monthly=canon_m,
        canonical_species_annual=canon_sa,
        canonical_species_monthly=canon_sm,
        canonical_reference=canon_ref,
        derived_annual=derived,
    )


# ---------------------------------------------------------------------------
# Reporting helpers
# ---------------------------------------------------------------------------

_BANNER = "=" * 60


def _format_period_coverage(
    annual: Iterable[AnnualValue],
    monthly: Iterable[MonthlyValue],
) -> tuple[str, str]:
    """Return (annual_periods, monthly_periods) one-line summaries."""
    a_years = sorted({a.year for a in annual})
    m_keys = sorted({(m.year, m.month) for m in monthly})

    annual_str = ", ".join(str(y) for y in a_years) if a_years else "—"
    if not m_keys:
        monthly_str = "—"
    else:
        first = m_keys[0]
        last = m_keys[-1]
        monthly_str = (
            f"{first[0]}-{first[1]:02d} … {last[0]}-{last[1]:02d}"
            f"  ({len(m_keys)} cells)"
        )
    return annual_str, monthly_str


SUMMARY_WARN_LIMIT = 5


def _format_warnings(warnings: list[str], limit: int = SUMMARY_WARN_LIMIT) -> list[str]:
    if not warnings:
        return ["  (none)"]
    head = warnings[:limit]
    out = [f"  - {w}" for w in head]
    if len(warnings) > limit:
        out.append(f"  ... ({len(warnings) - limit} more — use -v to see all)")
    return out


def print_summary(xlsx_path: Path, pipeline: PipelineOutput) -> None:
    raw = pipeline.raw_result
    annual_str, monthly_str = _format_period_coverage(
        pipeline.canonical_annual, pipeline.canonical_monthly
    )
    derived_count = len(pipeline.derived_annual)
    derived_metrics = sorted({d.metric_code for d in pipeline.derived_annual})
    derived_detail = (
        ", ".join(f"{m} × {sum(1 for d in pipeline.derived_annual if d.metric_code == m)}"
                  for m in derived_metrics)
        if derived_metrics else "(none)"
    )

    lines: list[str] = []
    lines.append(_BANNER)
    lines.append(f"WRITEBACK PLAN — {xlsx_path.name}")
    lines.append(_BANNER)
    lines.append(f"Parser:           {pipeline.parser_name}")
    lines.append(f"Source file:      {xlsx_path}")
    ref_count = len(pipeline.canonical_reference)
    ref_categories = sorted({r.category for r in pipeline.canonical_reference})
    ref_detail = ", ".join(ref_categories) if ref_categories else "(none)"

    lines.append("")
    lines.append("Counts (canonical, deduplicated):")
    lines.append(f"  Annual values:    {len(pipeline.canonical_annual)}")
    lines.append(f"  Monthly values:   {len(pipeline.canonical_monthly)}")
    lines.append(f"  Species annual:   {len(pipeline.canonical_species_annual)}")
    lines.append(f"  Species monthly:  {len(pipeline.canonical_species_monthly)}")
    lines.append(f"  Derived metrics:  {derived_count}  ({derived_detail})")
    lines.append(f"  Reference texts:  {ref_count}  ({ref_detail})")
    lines.append("")
    lines.append("Period coverage:")
    lines.append(f"  Annual:   {annual_str}")
    lines.append(f"  Monthly:  {monthly_str}")
    lines.append("")
    lines.append(f"Warnings: {len(raw.warnings)}")
    lines.extend(_format_warnings(raw.warnings))
    lines.append("")
    lines.append(f"Errors:   {len(raw.errors)}")
    if raw.errors:
        lines.extend(f"  - {e}" for e in raw.errors)
    lines.append(_BANNER)
    print("\n".join(lines))


def print_outcome(outcome: WriteResult) -> None:
    print(_BANNER)
    print("WRITE RESULT")
    print(_BANNER)
    print(f"Batch id:           {outcome.batch_id}")
    print(f"Rows to revisions:  {outcome.rows_to_revisions}")
    print(f"Rows to canonical:  {outcome.rows_to_canonical}")
    print(f"Rows unchanged:     {outcome.rows_unchanged}")
    print(f"Rows superseded:    {outcome.rows_superseded}")
    if outcome.warnings:
        print("Warnings:")
        for w in outcome.warnings:
            print(f"  - {w}")
    if outcome.errors:
        print("Errors:")
        for e in outcome.errors:
            print(f"  - {e}")
    print(_BANNER)


# ---------------------------------------------------------------------------
# Batch construction
# ---------------------------------------------------------------------------

def _build_batch(xlsx_path: Path, pipeline: PipelineOutput) -> WriteBatch:
    """Pack canonical+derived facts into a single ``WriteBatch``."""
    from etl.db.batch import build_batch_from_canonical

    # Derived annual facts share the same canonical contract as raw ones —
    # they enter the batch alongside, distinguished only by source_priority=99
    # and source_file="(derived)" set in compute_derived_annual.
    merged_annual = list(pipeline.canonical_annual) + list(pipeline.derived_annual)

    # Vintage date for the batch as a whole = latest vintage seen in any
    # canonical fact (so retries with newer-vintage data update correctly).
    all_vintages = [
        f.vintage_date for f in merged_annual
    ] + [
        f.vintage_date for f in pipeline.canonical_monthly
    ] + [
        f.vintage_date for f in pipeline.canonical_species_annual
    ] + [
        f.vintage_date for f in pipeline.canonical_species_monthly
    ] + [
        f.vintage_date for f in pipeline.canonical_reference
    ]
    vintage = max(all_vintages) if all_vintages else datetime.now().astimezone()

    return build_batch_from_canonical(
        source_file=str(xlsx_path),
        vintage_date=vintage,
        annual=merged_annual,
        monthly=pipeline.canonical_monthly,
        species_annual=pipeline.canonical_species_annual,
        species_monthly=pipeline.canonical_species_monthly,
        reference=pipeline.canonical_reference,
    )


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="etl.cli.writeback",
        description="Parse one Excel file and push canonical facts to on-prem Postgres.",
    )
    p.add_argument("xlsx_path", type=Path, help="Path to the Excel file to ingest.")
    p.add_argument(
        "--commit",
        action="store_true",
        help="Apply writes to DB. Without this flag the command is a dry-run.",
    )
    p.add_argument(
        "--config-env",
        type=Path,
        default=Path(".env"),
        help="Path to the .env file with SSH/DB credentials. Default: ./.env",
    )
    p.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Print extra information (full warning list, traceback on failure).",
    )
    return p


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)

    # Windows defaults stdout to cp1251, which can't encode Ukrainian
    # punctuation in our Excel filenames / Cyrillic content snippets.
    # Force utf-8 only when stdout reports another codec (pattern reused
    # from admin_report.py / diff.py).
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]

    if not args.xlsx_path.is_file():
        print(f"ERROR: file not found: {args.xlsx_path}", file=sys.stderr)
        return 2

    # ------------------------------------------------------------------
    # Phase 1: parse + canonical + derived (always runs)
    # ------------------------------------------------------------------
    try:
        pipeline = run_full_pipeline(args.xlsx_path)
    except ValueError as exc:
        # detect_parser failure
        print(f"ERROR (parser auto-detect): {exc}", file=sys.stderr)
        return 2

    print_summary(args.xlsx_path, pipeline)

    if pipeline.raw_result.errors:
        print(
            f"\nParser reported {len(pipeline.raw_result.errors)} error(s); "
            f"refusing to continue.",
            file=sys.stderr,
        )
        return 3

    if args.verbose and len(pipeline.raw_result.warnings) > SUMMARY_WARN_LIMIT:
        # Top 5 already shown by print_summary; continue from #6 to avoid duplication.
        rest = pipeline.raw_result.warnings[SUMMARY_WARN_LIMIT:]
        total = len(pipeline.raw_result.warnings)
        print(f"\nFull warning list ({total} total):")
        for i, w in enumerate(rest, start=SUMMARY_WARN_LIMIT + 1):
            print(f"  {i}. {w}")

    # ------------------------------------------------------------------
    # Phase 2: dry-run vs commit
    # ------------------------------------------------------------------
    if not args.commit:
        print("\n[DRY-RUN] No DB writes attempted. Use --commit to apply.")
        return 0

    print("\nConnecting to DB through SSH tunnel...")
    try:
        from etl.db.code_mapper import UnknownMetricError
        from etl.db.connection import open_connection
        from etl.db.postgres import PostgresRepository
    except ImportError as exc:
        print(f"ERROR: missing dependency for --commit path: {exc}", file=sys.stderr)
        return 5

    try:
        with open_connection(args.config_env) as conn:
            repo = PostgresRepository(conn)
            batch = _build_batch(args.xlsx_path, pipeline)
            outcome = repo.write_batch(batch)
    except FileNotFoundError as exc:
        # .env missing or unreadable
        print(f"ERROR (missing .env): {exc}", file=sys.stderr)
        return 4
    except UnknownMetricError as exc:
        # code_mapper missing entry OR DB has no indicators row for this code
        print(f"ERROR (unknown metric_code): {exc}", file=sys.stderr)
        return 6
    except KeyError as exc:
        # load_env validation — required env key missing
        print(f"ERROR (config): {exc}", file=sys.stderr)
        return 4
    except Exception as exc:  # noqa: BLE001 — surface DB transport failures cleanly
        if args.verbose:
            import traceback
            traceback.print_exc()
        print(f"ERROR (DB): {type(exc).__name__}: {exc}", file=sys.stderr)
        return 5

    print_outcome(outcome)
    return 0


if __name__ == "__main__":
    sys.exit(main())
